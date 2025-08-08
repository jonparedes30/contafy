from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from empresa.models import MateriaPrima, ProductoManufacturado, OrdenProduccion, ConsumoMateriaPrima
from datetime import datetime
import io

@login_required
def exportar_excel_materias_primas(request):
    empresa = request.user.empresa
    materias = MateriaPrima.objects.filter(empresa=empresa)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Materias Primas"
    
    headers = ['Código', 'Nombre', 'Unidad', 'Precio Unitario', 'Stock Actual', 'Stock Mínimo']
    ws.append(headers)
    
    for materia in materias:
        ws.append([
            materia.codigo,
            materia.nombre,
            materia.get_unidad_medida_display(),
            float(materia.precio_unitario),
            float(materia.stock_actual),
            float(materia.stock_minimo)
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="materias_primas_{empresa.nombre}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
def exportar_excel_productos_manufacturados(request):
    empresa = request.user.empresa
    productos = ProductoManufacturado.objects.filter(empresa=empresa)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos Manufacturados"
    
    headers = ['Código', 'Nombre', 'Precio Venta', 'Precio Costo', 'Stock Actual', 'Tiempo Producción (min)']
    ws.append(headers)
    
    for producto in productos:
        ws.append([
            producto.codigo,
            producto.nombre,
            float(producto.precio_venta),
            float(producto.precio_costo),
            producto.stock_actual,
            producto.tiempo_produccion
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="productos_manufacturados_{empresa.nombre}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
def exportar_excel_ordenes_produccion(request):
    empresa = request.user.empresa
    ordenes = OrdenProduccion.objects.filter(empresa=empresa)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Órdenes de Producción"
    
    headers = ['Número Orden', 'Producto', 'Cantidad Solicitada', 'Cantidad Producida', 'Estado', 'Fecha Creación']
    ws.append(headers)
    
    for orden in ordenes:
        ws.append([
            orden.numero_orden,
            orden.producto.nombre,
            orden.cantidad_solicitada,
            orden.cantidad_producida,
            orden.get_estado_display(),
            orden.creado_en.strftime('%d/%m/%Y')
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="ordenes_produccion_{empresa.nombre}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
def exportar_pdf_manufactura_completo(request):
    empresa = request.user.empresa
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []
    
    # Título
    title = Paragraph(f"Reporte Completo de Manufactura - {empresa.nombre}", styles['Title'])
    elements.append(title)
    
    # Materias Primas
    elements.append(Paragraph("Materias Primas", styles['Heading2']))
    materias = MateriaPrima.objects.filter(empresa=empresa)
    data = [['Código', 'Nombre', 'Stock', 'Precio']]
    for materia in materias:
        data.append([materia.codigo, materia.nombre, str(materia.stock_actual), f"${materia.precio_unitario}"])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    
    # Productos Manufacturados
    elements.append(Paragraph("Productos Manufacturados", styles['Heading2']))
    productos = ProductoManufacturado.objects.filter(empresa=empresa)
    data = [['Código', 'Nombre', 'Stock', 'Precio Venta']]
    for producto in productos:
        data.append([producto.codigo, producto.nombre, str(producto.stock_actual), f"${producto.precio_venta}"])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_manufactura_{empresa.nombre}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response